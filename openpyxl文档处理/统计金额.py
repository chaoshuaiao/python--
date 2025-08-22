# from datetime import datetime
from openpyxl.formula.translate import Translator
from openpyxl import  load_workbook


#传入时间戳公式,将零时区转换成东八区 
# ws2["d2"] = "=(c2+8*3600)/86400+70*365+19"
# cell_range =    ws2.iter_rows(3,len,4,4)
# for cell in cell_range:    
#     cell[0].value = Translator(formula="=(c2+8*3600)/86400+70*365+19",origin ="d2").translate_formula(cell[0].coordinate)
# for cell in ws2["d"]:
#     cell.number_format =    numbers.FORMAT_DATE_DATETIME



wb = load_workbook(r"D:\桌面\汇率.xlsx")
wb2 = load_workbook(r"D:\桌面\salesreport_202409.xlsx")

ws =  wb.active #ws =  wb.worksheets[0] or ws = wb['Sheet1']
ws2= wb2.active

#将汇率表中数据复制到需要整理的表中
cells = ws.iter_rows(1,48,1,2)
for cell in cells:
    ws2.cell(cell[0].row,18,cell[0].value)
    ws2.cell(cell[0].row,19,cell[1].value)


len =   len(ws2["a"]) #获取最大行数 

#计算相应汇率放到T列
ws2.cell(1,20,"汇率")
ws2["t2"] = "=VLOOKUP(j2,R:S,2,0)"
cell_huilv =    ws2.iter_rows(3,len,20,20)
for cell in cell_huilv:
    cell[0].value = Translator(formula="=VLOOKUP(j2,R:S,2,0)",origin ="t2").translate_formula(cell[0].coordinate)

#根据汇率计算相应的内购金额，放到U列
ws2.cell(1,21,"金额")
ws2["u2"] = "=t2*m2"
cell_pay = ws2.iter_rows(3,len,21,21)
for cell in cell_pay:
    cell[0].value = Translator(formula="=t2*m2",origin ="u2").translate_formula(cell[0].coordinate)




        


wb2.save(r"D:\桌面\salesreport_202409.xlsx")
print("已完成")