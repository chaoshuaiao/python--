import openpyxl as pxl


book = pxl.load_workbook(r"D:\桌面\afsd\adfa.xlsx",data_only=True) #book就是整个Excel文件
sheet = book.worksheets[0] #取第0张工作表 同 sheet = book.active 同 sheet = book["sheet1"] "sheet"是第一个工作表名
    #book.worksheets 代表所有工作表，可以用for循环进行遍历
# print(sheet.title)#取出工作表名
# print(sheet.min_row, sheet.max_row) #取出最小有效行号，最大有效行号
# print(sheet.min_column, sheet.max_column) #取出最小有效列号，最大有效列号
# for row in sheet.rows:                    #按行遍历整个工作表，从第一行到max_row行
#     for cell in row:                      #遍历一行的单元格，cell是一个单元格
#             print(cell.value,end="\t")                                        #取出单元格的值，空为None
# for cell in sheet["adfa1"]:               #                                           #按列遍历
#     print(cell.value)                                            #取出单元格的值，空为None
# for cell in sheet[2]:                                                #按行遍历
#     print(type(cell.value), cell.coordinate ,cell.col_idx , cell.number_format) #单元格属性：分别是单元格类型（int，str，datetime），单元格坐标，单元格列号，单元格的显示格式（百分比，分数之类）
# print(pxl.utils.get_column_letter(1)) #根据列名求取列名
# print(pxl.utils.column_index_from_string("A"))#根据列名求取列号
# for cell in sheet[2]:
#     print(cell.value)
cell = sheet.cell(row=2,column=2) #取出第2行第2列的单元格
cell.value = "-12.5"

if cell.value.isdigit():
    pass
else:
    try:
        cell.value = float(cell.value)
    except:
        cell.value = 0

print(cell.value,type(cell.value))

