from bs4 import BeautifulSoup
from matplotlib.table import Cell
from openpyxl import Workbook, load_workbook
import requests

wb =    load_workbook("D:\桌面\常用文件\汇率.xlsx")

ws =    wb.active
for i in range(3,47):
    cell =  ws.cell(i,1)
    target_currency = cell.value
    # print(target_currency)
    url = f"https://wise.com/zh-cn/currency-converter/{target_currency}-to-USD-rate?amount=1"  # 替换为实际的URL  
    
    response = requests.get(url)  
    response.raise_for_status()  # 如果请求失败，抛出异常

    soup = BeautifulSoup(response.text,'lxml')
        #     # 假设汇率信息在一个id为'exchange-rate-value'的元素中  
    obj = soup.find('span',class_='text-success').get_text()
        #     # 假设你有一个API或网页URL可以获取汇率  
    
    cell2 = ws.cell(i,2)
    cell2.value = obj
wb.save("D:\桌面\汇率2.xlsx")


    